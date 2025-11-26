# coding=utf-8
__version__ = "1.0.1"

''' Copyright (c) 2025 Rolling Loayza

    Permission is hereby granted, free of charge, to any person obtaining a copy
    of this software and associated documentation files (the "Software"), to deal
    in the Software without restriction, including without limitation the rights
    to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the Software is
    furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in all
    copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
    OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
    SOFTWARE.
'''

import os
import sys
import subprocess
import locale
import platform
import configparser
import math
import time
import csv
from checkcode import randomfile, getfile2json

# Settings
url = "https://raw.githubusercontent.com/dde-siei/checkcode/main/versioning.json"
scriptname = os.environ['scriptname']
shab = os.environ['shab']
# Get Oficial Version
jsonshaa = getfile2json(url)
shaa = jsonshaa[scriptname][__version__]
# Show Header
print('-'*64)
print('{:<7}: {}'.format('Script',scriptname))
print('{:<7}: {}'.format('Version',__version__))
print('Copyright(c) 2025 SERNANP-DDE-SIEI. All rights reserved.')
print('-'*64)
if shab == shaa:
    msginteg='Pass'
else:
    msginteg ='Not Pass'

print('|{:^62}|'.format('Check Integrity Script'))
print(shab)
print(shaa)
print('*{:^62}*'.format(msginteg))
print(f'{'-'*64}\n')

CONFIG_FILE = 'config.ini'
config = configparser.ConfigParser()
exportpath = os.path.join(os.environ['USERPROFILE'],'Downloads')

if not os.path.exists(exportpath):
    exportpath = os.path.join(os.environ['USERPROFILE'],'Descargas')

if not os.path.exists(CONFIG_FILE):
    print(f"Creating File {CONFIG_FILE}...\n")
    # Definir la configuración por defecto
    config['reportSFL'] = {
        'source': os.getcwd(),
        'root': r'\\10.10.11.21\new-archivod$\SINANPE',
        'folder': r'Sanea_Fis_Leg\Demarc_Fis',
        'export': exportpath
    }

    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        config.write(f)
else:
    # Lectura de configuración existente
    config.read(CONFIG_FILE)

# Check host function
def ping_host(host, attempts=2, timeout=1000):
    param_count = "-n" if platform.system().lower() == "windows" else "-c"
    param_timeout = "-w" if platform.system().lower() == "windows" else "-W"
    cmd = ["ping", param_count, str(attempts), param_timeout, str(timeout), host]

    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return result.returncode == 0
    except Exception:
        return False

testmsg = [
    'Getting User',
    'Setting Root',
    'Setting Folder',
    'Connecting Host',
    'Connecting SADT'
]

# Proceso paso a paso con animación
for i in range(1, 6):
    sys.stdout.write("\r{:>15}|{}".format(testmsg[i-1], "." * i))
    match i:
        case 1:
            serviuser = os.environ['USERNAME']
        case 2:
            inFile = config['reportSFL']['root']
            host = inFile.split("\\")[2] 
        case 3:
            acces = config['reportSFL']['folder']
        case 4:
            if not ping_host(host):
                print(f"\nFail Connection {host}")
                sys.exit(1)
        case 5:
            try:
                os.listdir(inFile)
            except Exception:
                print('\rFail Connection SADT !')
                sys.exit(1)
    time.sleep(1)
    sys.stdout.flush()

print('\rConnected to SADT')
print(f'\n[Login]: {serviuser}')
a = os.listdir(inFile)

d = {
    'PN': 'Parque Nacional',
    'RN': 'Reserva Nacional',
    'SN': 'Santuario Nacional',
    'SH': 'Santuarios Históricos',
    'RPJ': 'Reservas Paisajísticas',
    'RVS': 'Refugios de Vida Silvestre',
    'RC': 'Reservas Comunales',
    'BP': 'Bosques de Protección',
    'CC': 'Cotos de Caza'
    }

print('-'*64)
print('|{:^62}|'.format(''))
print('|{:^62}|'.format('Sistema Archivo Digital Técnico (SADT)'))
print('|{:^62}|'.format(''))
print('-'*64)

time.sleep(1)
print('\n[CATEGORIA]:')
for n ,i in enumerate(a , start=1):
    if i in d:
        print('{:>5} {}'.format(n,d[i]))
    else:
        pass
time.sleep(1) 
cate = int(input('\nEscribe el número de consulta de: '))-1

time.sleep(1)
print('\n[ANP]:')

dl = os.scandir(os.path.join(inFile, a[cate]))
anp_list = []
for n, j in enumerate(dl ,start=1):
    anp_list.append(j.name)
    print('{:>5} {}'.format(n,j.name))
time.sleep(1) 
anp = int(input('\nEscribe el número de consulta de: '))-1
dd = os.path.join(inFile, a[cate],anp_list[anp], acces)

# Set spreadsheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font , Alignment

wb = Workbook()
ws = wb.active
wcolumn = [20,35,35,35,15] 
headers = [
    'Carpeta Principal',
    'Nivel 1',
    'Nivel 2',
    'Nivel 3',
    'Estado'
    ]

ws.append(headers)

# Define styles
fill = PatternFill(start_color='FF808080',  
                   end_color='FF808080',
                   fill_type='solid')

fill2 = PatternFill(start_color='FFCCCCCC',  
                   end_color='FFCCCCCC',
                   fill_type='solid')

fill3 = PatternFill(start_color='FFEEEEEE',
                    end_color='FFEEEEEE',
                    fill_type='solid')

whitefill = PatternFill(start_color='FFFFFFFF',
                    end_color='FFFFFFFF',
                    fill_type='solid')

font = Font(name='Arial', color='FFFFFFFF', bold=True)

# Apply the fill to the cell
for col in range(1, 6):  # Columns A–E
    cell = ws.cell(row=1, column=col)
    cell.fill = fill
    cell.font = font

header_align = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.alignment = header_align

def fillwhite(xrow,exc, side):
    ycolu = [i+1 for i in range(0,len(headers))]
    if side == 'r':
        levelc = ycolu[exc:]
    elif side == 'l':
        levelc = ycolu[:exc]
    elif side == 'b':
        ycolu.remove(exc)
        levelc = ycolu
    else:
        print( 'No Option, only r or l')
            
    for ih in levelc:
        cell = ws.cell(row=xrow, column=ih)
        cell.fill = whitefill

print(f'\n{acces}\\')
levelm =[]
rowm =[2]
row =0
for root, dirs, files in os.walk(dd):
    row +=1
    if 'Thumbs.db' in files:
        files.remove('Thumbs.db')
    else:
        pass      
    level = root.replace(dd, "").count(os.sep)
    levelm.append(level)

    if level == 0:
        levelm[0] = 1    
    else:        
        if levelm[-1] - levelm[-2] == 1:
            row -=1
        else:
            fillwhite(row,level, 'l')
            
        indent = "│   " * (level - 1) + ("├── " if level > 0 else "")
        rowm.append(row)
        time.sleep(0.1)
        print(f'{indent} {os.path.basename(root)}')
        fillwhite(row,level,'r')
        cell = ws.cell(row=row,column=level)
        cell.value = os.path.basename(root)
        cell.fill = fill2
        cell.alignment = Alignment(vertical="top")
       
    if len(files) >0: 
        level +=1
        row -=1
        for n,file in enumerate(files):
            row +=1 
            levelm.append(level)
            rowm.append(row)
            file_indent = "│   " * level + ("├── " if n < len(files)-1 else "└── ")
            time.sleep(0.1)
            print(f'{file_indent} {file}')
            if n < 1:
                fillwhite(row,level, 'r')
            else:                
                fillwhite(row,level, 'b')
            cell = ws.cell(row=row,column=level)
            cell.value = file
            cell.fill = fill3
            ihei = math.ceil((len(file)*1.2)/wcolumn[level-1])
            ws.row_dimensions[row].height = 15 * ihei            
            cell.alignment = Alignment(vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 30

fechar =time.strftime('%Y/%m/%d %H:%M Horas', time.localtime())

for icol in range(0,len(headers)):
    ws.column_dimensions[chr(65+icol)].width = wcolumn[icol]

cell = ws.cell(row=row+1,column=1)
cell.value = (
    "Fuente: Sistema de Archivo Digital Técnico (SADT), fecha: {}\n"
    "Elaboración: SERNANP-DDE-SIEI"
).format(fechar)
cell.font = Font(name='Arial', size= 9,color='FF000000', italic=True,bold=False)
cell.alignment = Alignment(vertical="center",wrap_text=True)

ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=5)
ws.row_dimensions[row+1].height = 40

print('\n[ RESUMEN DE REPORTE ]:\n')
time.sleep(1)
print('{:<10}: {}'.format('Categoría',d[a[cate]]))
print('{:<10}: {}'.format('ANP',anp_list[anp]))

os.chdir(config['reportSFL']['export'])

v = 1
while True:
    exname = f'reporte_{anp_list[anp]}_v{v}.xlsx'
    if not os.path.exists(exname):
       break
    v +=1

wb.save(exname)

xlxsname = os.path.join(os.getcwd(),exname)
time.sleep(1)
print('\nRuta de Reporte:')
time.sleep(2)
print('-'*(len(xlxsname)+22))
print('|{0:^{1}}|'.format(xlxsname, (len(xlxsname)+20)))
print('-'*(len(xlxsname)+22))

os.chdir(config['reportSFL']['source'])

lastversions = [key for key in jsonshaa[scriptname]]

def tempfup(lines):
    with open("tmp.dat","w") as f:
        f.write(lines)

flag = ''
time.sleep(2)
if __version__ == lastversions[-1]:
    if msginteg == 'Not Pass':
        print(f'*Script: {scriptname}.py Might Be Corrupted')
    else:
        print(f'*Script: {scriptname}.py Run Properly')
else:
    print('*Nueva Versión Diponible: {}'.format(lastversions[-1]))
    flag +='10'

if len(flag) > 0: 
    tempfup(flag)